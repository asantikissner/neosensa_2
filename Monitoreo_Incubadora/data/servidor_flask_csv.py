# servidor_flask_csv.py (CSV version con claves ordenadas)
from flask import Flask, jsonify
from flask_cors import CORS
from statistics import mean, median, stdev
import os, csv, glob, re
from flask import send_file
import io
from datetime import datetime

app = Flask(__name__)
CORS(app)

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "scripts", "data"))
MAX_ROWS = 720  # hasta 720 muestras

def calcular_estadisticas(valores):
    nums = []
    for v in valores:
        try:
            nums.append(float(v))
        except:
            pass
    if len(nums) < 2:
        return {"media": None, "mediana": None, "min": None, "max": None, "desvio": None}
    return {
        "media":   round(mean(nums), 2),
        "mediana": round(median(nums), 2),
        "min":     round(min(nums), 2),
        "max":     round(max(nums), 2),
        "desvio":  round(stdev(nums), 2)
    }

def read_csv_rows(path, max_rows=MAX_ROWS):
    rows = []
    if not os.path.exists(path):
        return rows
    with open(path, newline='', encoding='utf-8') as f:
        reader = list(csv.DictReader(f))
        rows = reader[:max_rows]  # primeras max_rows (asumo más recientes arriba)
    return rows

@app.route("/datos")
def obtener_datos_csv():
    resultado_final = {}
    os.makedirs(DATA_DIR, exist_ok=True)
    csv_paths = sorted(glob.glob(os.path.join(DATA_DIR, "incubadora_*.csv")))

    for path in csv_paths:
        try:
            # extraer número real del archivo
            filename = os.path.basename(path)  # ej: incubadora_7.csv
            match = re.search(r"incubadora_(\d+)\.csv", filename)
            if not match:
                continue
            esp_num = int(match.group(1))
            esp_id = f"incubadora_{esp_num}"

            filas = read_csv_rows(path, MAX_ROWS)
            if not filas:
                continue

            # estructuras de salida
            datos_T, historial_T = [], []
            datos_H, historial_H = [], []
            datos_I, historial_I = [], []
            datos_O, historial_O = [], []
            salt_T = set(); salt_H = set(); salt_I = set(); salt_O = set()

            for i, fila in enumerate(filas):
                fecha = fila.get("fecha")

                # TEMPERATURA
                if i not in salt_T:
                    temp = fila.get("temperatura")
                    ley_t = fila.get("ley_T", "-")
                    if ley_t == "-":
                        cont = 0; suma = 0.0; j = i
                        while j < len(filas) and filas[j].get("ley_T", "-") == "-" and cont < 6:
                            try: suma += float(filas[j].get("temperatura", 0))
                            except: pass
                            salt_T.add(j); cont += 1; j += 1
                        if cont > 0:
                            medio = i + cont//2
                            fecha_med = filas[medio].get("fecha")
                            datos_T.append({"fecha_grafico": fecha_med, "temperatura_grafico": round(suma/cont, 4)})
                    else:
                        try: valor = float(temp)
                        except: valor = None
                        if valor is not None:
                            historial_T.append({"valor": valor, "tipo": ley_t, "fecha": fecha})
                            datos_T.append({"fecha_grafico": fecha, "temperatura_grafico": valor})

                # HUMEDAD
                if i not in salt_H:
                    hum = fila.get("humedad")
                    ley_h = fila.get("ley_H", "-")
                    if ley_h == "-":
                        cont = 0; suma = 0.0; j = i
                        while j < len(filas) and filas[j].get("ley_H", "-") == "-" and cont < 6:
                            try: suma += float(filas[j].get("humedad", 0))
                            except: pass
                            salt_H.add(j); cont += 1; j += 1
                        if cont > 0:
                            medio = i + cont//2
                            fecha_med = filas[medio].get("fecha")
                            datos_H.append({"fecha_grafico": fecha_med, "humedad_grafico": round(suma/cont, 4)})
                    else:
                        try: valor = float(hum)
                        except: valor = None
                        if valor is not None:
                            historial_H.append({"valor": valor, "tipo": ley_h, "fecha": fecha})
                            datos_H.append({"fecha_grafico": fecha, "humedad_grafico": valor})

                # ILUMINANCIA
                if i not in salt_I:
                    lux = fila.get("iluminancia")
                    ley_i = fila.get("ley_I", "-")
                    if ley_i == "-":
                        cont = 0; suma = 0.0; j = i
                        while j < len(filas) and filas[j].get("ley_I", "-") == "-" and cont < 6:
                            try: suma += float(filas[j].get("iluminancia", 0))
                            except: pass
                            salt_I.add(j); cont += 1; j += 1
                        if cont > 0:
                            medio = i + cont//2
                            fecha_med = filas[medio].get("fecha")
                            datos_I.append({"fecha_grafico": fecha_med, "iluminancia_grafico": round(suma/cont, 4)})
                    else:
                        try: valor = float(lux)
                        except: valor = None
                        if valor is not None:
                            historial_I.append({"valor": valor, "tipo": ley_i, "fecha": fecha})
                            datos_I.append({"fecha_grafico": fecha, "iluminancia_grafico": valor})

                # OXIGENO
                if i not in salt_O:
                    ox = fila.get("oxigeno")
                    ley_o = fila.get("ley_O", "-")
                    if ley_o == "-":
                        cont = 0; suma = 0.0; j = i
                        while j < len(filas) and filas[j].get("ley_O", "-") == "-" and cont < 6:
                            try: suma += float(filas[j].get("oxigeno", 0))
                            except: pass
                            salt_O.add(j); cont += 1; j += 1
                        if cont > 0:
                            medio = i + cont//2
                            fecha_med = filas[medio].get("fecha")
                            datos_O.append({"fecha_grafico": fecha_med, "oxigeno_grafico": round(suma/cont, 4)})
                    else:
                        try: valor = float(ox)
                        except: valor = None
                        if valor is not None:
                            historial_O.append({"valor": valor, "tipo": ley_o, "fecha": fecha})
                            datos_O.append({"fecha_grafico": fecha, "oxigeno_grafico": valor})

            # invertir para que vayan de más antiguo a más reciente
            datos_T.reverse(); datos_H.reverse(); datos_I.reverse(); datos_O.reverse()

            # valores actuales
            top = filas[0]
            try: Tact = float(top.get("temperatura"))
            except: Tact = None
            try: Hact = float(top.get("humedad"))
            except: Hact = None
            try: Iact = float(top.get("iluminancia"))
            except: Iact = None
            try: Oact = float(top.get("oxigeno"))
            except: Oact = None

            estad = {
                "temperatura": calcular_estadisticas([r.get("temperatura") for r in filas]),
                "humedad": calcular_estadisticas([r.get("humedad") for r in filas]),
                "iluminancia": calcular_estadisticas([r.get("iluminancia") for r in filas]),
                "oxigeno": calcular_estadisticas([r.get("oxigeno") for r in filas])
            }

            ultimo_timestamp = None
            if filas:
                try:
                    ultimo_timestamp = filas[0].get("fecha")  # asumo formato datetime string
                except:
                    pass

            resultado_final[esp_id] = {
                "T_actual": Tact,
                "H_actual": Hact,
                "I_actual": Iact,
                "O_actual": Oact,
                "datos_T": datos_T,
                "datos_H": datos_H,
                "datos_I": datos_I,
                "datos_O": datos_O,
                "historial_T": historial_T,
                "historial_H": historial_H,
                "historial_I": historial_I,
                "historial_O": historial_O,
                "estadisticas": estad,
                "ultimo_timestamp": ultimo_timestamp   # 👈 agregado
            }

        except Exception as e:
            print(f"Error procesando {path}: {e}")

    # 👇 ordenar las claves numéricamente antes de devolver
    ordenado = dict(sorted(
        resultado_final.items(),
        key=lambda kv: int(kv[0].split("_")[1])
    ))

    return jsonify(ordenado)

@app.route("/download_xlsx")
def download_xlsx():
    """
    Genera un Excel en memoria donde cada CSV 'incubadora_N.csv' es una hoja.
    Intenta usar pandas; si no está instalado, hace fallback con openpyxl.
    """
    # buscar CSVs
    csv_dir = DATA_DIR
    csv_paths = sorted(glob.glob(os.path.join(csv_dir, "incubadora_*.csv")))
    if not csv_paths:
        return jsonify({"error": "no csv files found"}), 404

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"incubadoras_{ts}.xlsx"

    # intentar con pandas (más simple y preserva headers)
    try:
        import pandas as pd
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for path in csv_paths:
                sheet_name = os.path.splitext(os.path.basename(path))[0]
                # sheets names can't exceed 31 chars
                sheet_name = sheet_name[:31]
                try:
                    df = pd.read_csv(path, encoding='utf-8')
                except Exception:
                    # si falla leer con utf-8, intentar latin-1
                    df = pd.read_csv(path, encoding='latin-1')
                # escribir hoja
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        output.seek(0)
        return send_file(output,
                         as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e_pandas:
        # fallback: usar openpyxl directamente (sin pandas)
        try:
            from openpyxl import Workbook
            wb = Workbook()
            # eliminar sheet por defecto si lo tiene
            if wb.sheetnames:
                std = wb[wb.sheetnames[0]]
                wb.remove(std)
            for path in csv_paths:
                sheet_name = os.path.splitext(os.path.basename(path))[0][:31]
                ws = wb.create_sheet(title=sheet_name)
                # leer csv y volcar filas
                with open(path, newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for r_idx, row in enumerate(reader, start=1):
                        for c_idx, cell in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=cell)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(bio,
                             as_attachment=True,
                             download_name=filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e_openpyxl:
            # si todo falla, devolver error
            print("Error generando xlsx:", e_pandas, e_openpyxl)
            return jsonify({"error": "failed to generate xlsx"}), 500
        
if __name__ == "__main__":
    print("✅ Servidor Flask (CSV completo + ordenado) iniciado en http://localhost:5000/datos")
    app.run(host="0.0.0.0", port=5000, debug=True)
