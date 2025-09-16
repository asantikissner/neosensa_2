import threading
import time
import json
import random
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import os
from pathlib import Path
import paho.mqtt.client as mqtt

app = Flask(__name__)
CORS(app)  # permite todas las orígenes

# Valores por defecto
Tmin, Tmax, Ttol = 36, 38, 1
Hmin, Hmax, Htol = 85, 95, 1
Imin, Imax, Itol = 500, 2200, 1
Omin, Omax, Otol = 78, 91, 1

temperatura=37
humedad=90
sonido=22
iluminancia=1000
oxigeno=90
###############################################################################################

# Planilla excel para almacenamiento local de datos
workbook = load_workbook(filename="/Users/antonellasantikissner/Documents/PROYECTO/neosensa/Monitoreo_Incubadora/incubadora_iot.xlsx")
sheet=workbook.sheetnames
incubadora_1 = workbook[sheet[0]]
incubadora_2=workbook[sheet[1]]


# 📁 Archivo de límites persistentes
limites_path = Path("limites.json")

 # --- Cargar los límites para ese ESP ---
DEFAULTS = {
    "Tmin": 36, "Tmax": 38, "Ttol": 1,
    "Hmin": 85,    "Hmax": 95, "Htol": 1,
    "Imin": 500,   "Imax": 2200, "Itol": 1,
    "Omin": 78,    "Omax": 91, "Otol": 1,
}


###############################################################################################

# Funciones para simulacion realista
def simular_ruido(valor_anterior, max_variacion):
    return round(valor_anterior + random.uniform(-max_variacion, max_variacion), 2)

def limitar(valor, minimo, maximo):
    return min(max(valor, minimo), maximo)

def on_connect(client, userdata, flags, rc):
    print("🔌 Conectado al broker MQTT")
    client.subscribe("config/limites")
    client.subscribe("sensor/+/datos") 

def on_message(client, userdata, msg):
    if msg.topic.startswith("sensor/") and msg.topic.endswith("/datos"):
        try:
            data = json.loads(msg.payload.decode())
            esp = str(data.get("esp", "1"))
            print(f"Mensaje recibido en topic: {msg.topic}")
            print(f"Payload: {msg.payload.decode()}")
            print(f"📡 MQTT recibido de esp {esp}")
            print(json.dumps(data, indent=2))

            # Guardar en Excel
            if esp == "1":
                hoja = incubadora_1
            else:
                hoja = incubadora_2

            hoja.insert_rows(idx=3)
            hoja["B3"] = data.get("fecha")
            hoja["C3"] = data.get("temperatura")
            hoja["D3"] = data.get("ley_T")
            hoja["E3"] = data.get("humedad")
            hoja["F3"] = data.get("ley_H")
            hoja["G3"] = data.get("iluminancia")
            hoja["H3"] = data.get("ley_I")
            hoja["I3"] = data.get("oxigeno")
            hoja["J3"] = data.get("ley_O")
            workbook.save(filename="/Users/antonellasantikissner/Documents/PROYECTO/neosensa/Monitoreo_Incubadora/incubadora_iot.xlsx")

        except Exception as e:
            print(f"❌ Error al procesar mensaje MQTT: {e}")

    if msg.topic == "config/limites":
        data = json.loads(msg.payload.decode())
        esp = str(data.get("esp", "1"))

        # Cargá o inicializá el JSON completo
        all_limits = limites_path.exists() and json.loads(limites_path.read_text()) or {}
        prev = all_limits.get(esp, DEFAULTS)
        # Construí solo las claves de límites
        nuevos = { k: data.get(k, prev.get(k)) for k in DEFAULTS }
        all_limits[esp] = nuevos
        limites_path.write_text(json.dumps(all_limits, indent=2))
        print(f"📥 Nuevos límites para ESP {esp}: {nuevos}")

cliente_mqtt = mqtt.Client()
cliente_mqtt.on_connect = on_connect
cliente_mqtt.on_message = on_message
cliente_mqtt.connect("broker.emqx.io", 1883)
cliente_mqtt.loop_start()

@app.route("/limites", methods=["GET"])
def enviar_limites_actuales():
    esp = request.args.get("esp", "1")
    if limites_path.exists():
        all_limits = json.loads(limites_path.read_text())
        return jsonify(all_limits.get(esp, DEFAULTS)), 200
    return jsonify(DEFAULTS), 200


def simular_datos(incubadora_2):
    # Simular datos de sensores (incubadora 2)
        global temperatura, humedad, sonido, iluminancia, oxigeno, Tmin, Tmax, Hmin, Hmax, Imin, Imax, Omin, Omax
        global DEFAULTS

        temperatura = limitar(simular_ruido(temperatura, 2), Tmin, Tmax)
        humedad = limitar(simular_ruido(humedad, 2), Hmin, Hmax)
        iluminancia = limitar(simular_ruido(iluminancia, 50), Imin, Imax)
        oxigeno = limitar(simular_ruido(oxigeno, 1.5), Omin, Omax)
        esp = "2"

        # Simular alertas ocasionales (50% de probabilidades) s
        if random.random() < 0.5:  
            temperatura += random.choice([2,-2])
        if random.random() < 0.5:  
            humedad += random.choice([5,-5])
        if random.random() < 0.5:  
            iluminancia += random.choice([400,-400])
        if random.random() < 0.5:  
            oxigeno += random.choice([8,-8])

        if limites_path.exists():
            with open(limites_path, "r") as f:
                all_limits = json.load(f)
                limites = all_limits.get(esp, DEFAULTS)
        else:
            limites = DEFAULTS

        Tmin, Tmax, Ttol = limites["Tmin"], limites["Tmax"], limites["Ttol"]
        Hmin, Hmax, Htol = limites["Hmin"], limites["Hmax"], limites["Htol"]
        Imin, Imax, Itol = limites["Imin"], limites["Imax"], limites["Itol"]
        Omin, Omax, Otol = limites["Omin"], limites["Omax"], limites["Otol"]

                # Alarmas
        alarma_T = True if ((temperatura < (Tmin-Ttol)) or (temperatura > (Tmax+Ttol))) else False
        alarma_H = True if (humedad < (Hmin-Htol) or humedad > (Hmax+Htol)) else False
        alarma_I = True if ((iluminancia < (Imin-Itol)) or (iluminancia > (Imax+Itol))) else False
        alarma_O = True if ((oxigeno < (Omin-Otol)) or (oxigeno > (Omax+Otol))) else False

        # Leyendas
        ley_T = "alta" if (temperatura > Tmax and alarma_T==1) else "baja" if (temperatura < Tmin and alarma_T==1) else "-"
        ley_H = "alta" if (humedad > Hmax and alarma_H==1) else "baja" if (humedad < Hmin and alarma_H==1) else "-"
        ley_I = "alta" if (iluminancia > Imax and alarma_I==1) else "baja" if (iluminancia < Imin and alarma_I==1) else "-"
        ley_O = "alto" if (oxigeno > Omax and alarma_O==1) else "bajo" if (oxigeno < Omin and alarma_O==1) else "-"

        ###########################################################################
        fecha=datetime.now().isoformat(timespec='seconds')
        datos_simulados = {
            "fecha":fecha,
            "temperatura": round(temperatura, 2),
            "humedad": round(humedad, 2),
            "iluminancia": round(iluminancia, 2),
            "oxigeno": round(oxigeno, 2),
            "Tmax": Tmax, "Tmin":Tmin, "Ttol":Ttol,
            "Hmax":Hmax, "Hmin":Hmin, "Htol":Htol,
            "Imax":Imax, "Imin":Imin, "Itol":Itol,
            "Omax":Omax, "Omin":Omin, "Otol":Otol,
            "alarma_T": alarma_T,
            "alarma_H": alarma_H,
            "alarma_I": alarma_I,
            "alarma_O": alarma_O,
            "ley_T":ley_T,
            "ley_H":ley_H,
            "ley_I":ley_I,
            "ley_O":ley_O,
            "esp":2,
        }

        # Publicar por MQTT
        cliente_mqtt.publish("sensor/2/datos", json.dumps(datos_simulados))

def bucle_simulacion():
    inc_id = 2
    while True:
        simular_datos(incubadora_2)
        time.sleep(10)

if __name__ == "__main__":
    print("✅ Servidor Flask Simulacion iniciado en http://localhost:5001/datos")
    threading.Thread(target=bucle_simulacion, daemon=True).start()
    app.run(host="0.0.0.0", port=5001)
    
